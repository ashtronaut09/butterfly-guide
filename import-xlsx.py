#!/usr/bin/env python3
"""
Parse Butterfly Purchases.xlsx into structured JSON collection.

Input:  /Users/ashley/Downloads/Butterfly Purchases.xlsx (Sheet1)
Output: /Users/ashley/Documents/butterfly-guide/data/collection.json
        /Users/ashley/Documents/butterfly-guide/data/import-report.txt

Columns in spreadsheet (Row 1 = header):
  A: unused
  B: notes / free text
  C: English name with embedded sex, location, form/aberration, altitude, quantity
  D: Latin name with French location, condition codes, sex qualifiers
  E: Supplier ('username' Name, Address)
  F: Price (£, numeric or "collected"/"free"/"job lot", etc.)
  G: Date bought
  H: Date sent
  I: Date received/set
  J: Setting board
  K: Cat number
"""

import json
import logging
import os
import re
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────
INPUT_FILE = Path("/Users/ashley/Downloads/Butterfly Purchases.xlsx")
OUTPUT_DIR = Path("/Users/ashley/Documents/butterfly-guide/data")
OUTPUT_JSON = OUTPUT_DIR / "collection.json"
OUTPUT_REPORT = OUTPUT_DIR / "import-report.txt"
SHEET_NAME = "Sheet1"

# ── Constants ──────────────────────────────────────────────────────────────

# Sex symbols we look for
SEX_MALE = "\u2642"  # ♂
SEX_FEMALE = "\u2640"  # ♀

# French location patterns in col D (Latin name column) and their English mappings
FRENCH_LOCATION_PATTERNS = [
    (re.compile(r"d'Espagne", re.IGNORECASE), "Spain"),
    (re.compile(r"des\s+A[çc]ores", re.IGNORECASE), "Azores"),
    (re.compile(r"du\s+Chypre", re.IGNORECASE), "Cyprus"),
    (re.compile(r"de\s+Mad[eè]re", re.IGNORECASE), "Madeira"),
    (re.compile(r"d'Italie\s+du\s+Sud", re.IGNORECASE), "Southern Italy"),
    (re.compile(r"d'Italie", re.IGNORECASE), "Italy"),
    (re.compile(r"des\s+[iî]les\s+Canaries", re.IGNORECASE), "Canary Isles"),
    (re.compile(r"d'[AÀ]ragon", re.IGNORECASE), "Aragon Spain"),
]

# Description-prefix keywords in col C and D — when found, the keyword + following words
# are treated as description (form/aberration/subspecies info)
DESC_PREFIXES = re.compile(
    r"\b(aberration|abberation|aberation|aberration"
    r"|extreme\s+aberation"
    r"|form|subspecies|var\.|f\.)\b",
    re.IGNORECASE,
)

# "ab " (short for aberration) as a standalone word — careful not to match "about" etc.
AB_PREFIX = re.compile(r"\bab\b", re.IGNORECASE)

# Collector identity strings — if found in supplier, set collector
COLLECTOR_PATTERNS = re.compile(
    r"Ewen\s+Adamson|E\.D\.\s+Adamson|E\s+D\s+Adamson", re.IGNORECASE
)

# ══════════════════════════════════════════════════════════════════════════
#  DATE CONVERSION
# ══════════════════════════════════════════════════════════════════════════

def format_date(value):
    """Convert a date cell value to an ISO date string (YYYY-MM-DD) or None."""
    if value is None:
        return None

    # datetime objects → ISO string
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    # String dates — try common formats
    if isinstance(value, str):
        cleaned = value.strip()
        # Date + note like "05/11/2012 damaged missing antennae"
        date_match = re.match(r"(\d{1,2}/\d{1,2}/\d{2,4})", cleaned)
        if date_match:
            cleaned = date_match.group(1)
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(cleaned, fmt).strftime("%Y-%m-%d")
            except (ValueError, IndexError):
                continue
        # Not a parseable date — store in report but return as-is for logging
        return None

    # Integer — assume Excel serial date number
    if isinstance(value, (int, float)):
        try:
            from datetime import timedelta

            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(value))).strftime("%Y-%m-%d")
        except Exception:
            return None

    return None


# ══════════════════════════════════════════════════════════════════════════
#  SEX EXTRACTION (from English name column)
# ══════════════════════════════════════════════════════════════════════════

def extract_sex_and_quantity(text):
    """Extract sex, quantity info, and pair indicators from col C text.

    Returns (sex, quantity_descriptions, cleaned_text).

    Handles:
      - ♂ and ♀ symbols
      - ♂+♀, ♂♀ (mixed pairs)
      - "P" / "p" as standalone token meaning pair
      - "2 x ♂", "2x♂", "2♂", "2 x ♀" etc. (quantity + sex)
      - "2m", "3f" etc. where 1-2 digits precede m/f = quantity+sex
        (3-4 digit numbers + m are altitude, handled separately)
      - "2f 1m" etc. (mixed quantity)
    """
    sex = None
    qty_parts = []
    pair = False

    # ── 1. Detect sex from ♂/♀ symbols ──────────────────────────────────
    has_male = SEX_MALE in text
    has_female = SEX_FEMALE in text
    if has_male and has_female:
        sex = "♂♀"
    elif has_male:
        sex = "♂"
    elif has_female:
        sex = "♀"

    # ── 2. Structured quantity patterns (process BEFORE symbol removal) ─
    # Pattern: N x ♂, N x ♀ (with actual sex symbols)
    # This must be done before symbols are replaced to handle "2 x ♂"
    def replace_qty_with_symbol(m):
        nonlocal sex
        qty = m.group(1)
        symbol = m.group(2)
        indicator = "m" if symbol == SEX_MALE else "f"
        qty_parts.append(f"{qty} x {indicator}")
        if sex is None:
            sex = "♂" if symbol == SEX_MALE else "♀"
        return ""

    text = re.sub(
        r'(\d+)\s*x\s*([♂♀])',
        replace_qty_with_symbol,
        text,
    )

    # Pattern: N x m/f/p (text-based indicator)
    # IMPORTANT: \b word boundary prevents matching "f" as part of "from"
    def replace_qty_symbol(m):
        nonlocal sex
        qty = m.group(1)
        indicator = m.group(2).lower()
        qty_parts.append(f"{qty} x {indicator}")
        if indicator in ("m", "male") and sex is None:
            sex = "♂"
        elif indicator in ("f", "femelle", "female") and sex is None:
            sex = "♀"
        elif indicator in ("p", "pair") and sex is None:
            sex = "♂♀"
        return ""

    text = re.sub(
        r'(\d+)\s*x\s*(m(?:ale)?|f(?:emelle|emale)?|p(?:air)?)\b',
        replace_qty_symbol,
        text,
        flags=re.IGNORECASE,
    )

    # ── 3. Now replace sex symbols with space ───────────────────────────
    # (prevents words merging, e.g. "Copper♂from")
    text = text.replace(SEX_MALE, " ").replace(SEX_FEMALE, " ")
    # Also handle "♂+♀" text patterns (the + may remain)
    text = re.sub(r'\s*\+\s*(?=from|$)', '', text)

    # ── 4. Direct quantity patterns: N♂, N♀, Nm, Nf ────────────────────
    # BUT distinguish altitude: 3-4 digit + m = altitude (skip)
    # 1-2 digit + m/f = quantity+sex
    # (N♂ and N♀ already processed above as text, but handle "2m"/"2f")
    def replace_direct_qty(m):
        nonlocal sex
        digits = m.group(1)
        indicator = m.group(2).lower()
        # Only 1-2 digit numbers are quantity (3-4 digit + m is altitude)
        if len(digits) <= 2:
            if indicator == "m":
                sex = "♂" if sex is None else sex
                qty_parts.append(f"{digits} m")
            elif indicator == "f":
                sex = "♀" if sex is None else sex
                qty_parts.append(f"{digits} f")
            elif indicator == "p":
                sex = "♂♀" if sex is None else sex
                qty_parts.append(f"{digits} p")
            return ""
        return m.group(0)  # Don't remove (it's altitude — keep in text)

    text = re.sub(r'(\d+)\s*([mMfFpP])\b', replace_direct_qty, text)

    # ── 5. P/p as standalone pair indicator ─────────────────────────────
    # Replace token 'p' or 'P' or 'p.' or 'P.' when standalone
    # (processed AFTER "N x P" so those are already handled)
    def replace_pair(m):
        nonlocal pair, sex
        pair = True
        if sex is None:
            sex = "♂♀"
        return ""
    text = re.sub(r'(?<!\w)p\.?(?!\w)', replace_pair, text, flags=re.IGNORECASE)

    # Clean up extra whitespace from removals
    text = re.sub(r'\s+', ' ', text).strip()

    return sex, qty_parts, text, pair


# ══════════════════════════════════════════════════════════════════════════
#  LOCATION EXTRACTION
# ══════════════════════════════════════════════════════════════════════════

def extract_location(text):
    """Extract 'from X' location from col C text.

    Returns (location, cleaned_text).
    Handles "from X" where X is any text until end or period.
    Also handles standalone location words like just a place name
    that might appear without 'from' (e.g., "Bulgaria", "Chancelade France").
    """
    location = None

    # Primary pattern: "from X" (possibly with "the")
    m = re.search(r'\bfrom\s+(?:the\s+)?(.+?)(?:\.\s*)?$', text, re.IGNORECASE)
    if m:
        loc = m.group(1).strip().rstrip(".")
        if loc and not loc.isspace():
            location = loc
        text = text[: m.start()].strip()

    # Also handle "Taken X" pattern (used by collector entries)
    m = re.search(r'\bTaken\s+(.+?)(?:\.\s*)?$', text, re.IGNORECASE)
    if m and location is None:
        loc = m.group(1).strip().rstrip(".")
        if loc and not loc.isspace():
            location = loc
        text = text[: m.start()].strip()

    text = re.sub(r'\s+', ' ', text).strip()
    return location, text


def extract_location_from_latin(text):
    """Extract location from French patterns in Latin name column.

    Returns (location, remaining_cleaned_text).
    """
    location = None
    for pattern, eng_name in FRENCH_LOCATION_PATTERNS:
        m = pattern.search(text)
        if m:
            location = eng_name
            text = pattern.sub("", text)
            break

    text = re.sub(r'\s+', ' ', text).strip()
    return location, text


# ══════════════════════════════════════════════════════════════════════════
#  ALTITUDE EXTRACTION
# ══════════════════════════════════════════════════════════════════════════

def extract_altitude(text):
    """Extract altitude from text (3-4 digit number followed by m).

    Handles:
      - "Alt 1000m", "Alt. 1000m"
      - "2100m", "1900m"
      - "1000m - 2500m" (takes first value)

    Returns (altitude_m, cleaned_text).
    """
    altitude = None

    # Pattern: optional "Alt" or "Alt." prefix, then 3-4 digits, then "m"
    m = re.search(r'\b(?:Alt\.?\s*)?(\d{3,4})\s*m\b', text, re.IGNORECASE)
    if m:
        altitude = int(m.group(1))
        text = text[: m.start()] + text[m.end() :]

    text = re.sub(r'\s+', ' ', text).strip()
    return altitude, text


# ══════════════════════════════════════════════════════════════════════════
#  DESCRIPTION EXTRACTION
# ══════════════════════════════════════════════════════════════════════════

def extract_description(text):
    """Extract description fragments (form, aberration, condition, etc.) from text.

    Returns (description_parts, cleaned_text).
    """
    parts = []
    text2 = text

    # ── Pattern 1: Full-word aberration/form prefixes with following words ─
    # e.g. "aberration Albino", "form chlorodippe", "Extreme aberation"
    # Trailing word is optional (handles keywords at end of string)
    def grab_keyword_block(m):
        fragment = m.group(0).strip()
        if fragment:
            parts.append(fragment)
        return ""

    while True:
        m = re.search(
            r'\b(aberration|abberation|aberation|aberration'
            r'|extreme\s+aberation|form|subspecies)'
            r'(?:\s+\w+){0,3}',
            text2,
            re.IGNORECASE,
        )
        if m:
            parts.append(m.group(0).strip())
            text2 = text2[: m.start()] + text2[m.end() :]
        else:
            break

    # ── Pattern 2: "ab" as short for aberration (standalone word) ──────────
    # Grab up to 3 following words, or just "ab" at end of text
    while True:
        m = re.search(r'\b(ab)\b\s*(?:(?:\w+\s+){0,2}\w+)?', text2, re.IGNORECASE)
        if m:
            # Verify "ab" is standalone, not part of another word
            start = m.start()
            if start == 0 or not text2[start - 1].isalpha():
                captured = m.group(0).strip()
                # Only accept if it has following words, or if "ab" is the
                # only thing left (terminal aberration abbreviation)
                words = captured.split()
                if len(words) >= 2 or (len(words) == 1 and words[0].lower() == 'ab'):
                    parts.append(captured)
                    text2 = text2[: m.start()] + text2[m.end() :]
                else:
                    break
            else:
                break
        else:
            break

    # ── Pattern 3: "Var" or "var." (variety) ──────────────────────────────
    m = re.search(r'\bvar\.?\s*\w*', text2, re.IGNORECASE)
    if m:
        parts.append(m.group(0).strip())
        text2 = text2[: m.start()] + text2[m.end() :]

    # ── Pattern 4: "f." (form) standalone ─────────────────────────────────
    m = re.search(r'\bf\.\s*\w+', text2, re.IGNORECASE)
    if m:
        parts.append(m.group(0).strip())
        text2 = text2[: m.start()] + text2[m.end() :]

    text2 = re.sub(r'\s+', ' ', text2).strip()
    return parts, text2


# ══════════════════════════════════════════════════════════════════════════
#  LATIN NAME CLEANING
# ══════════════════════════════════════════════════════════════════════════

def clean_latin_name(text):
    """Clean the Latin name column, extracting extra info to description.

    Strips:
      - French location patterns (→ location field, not description)
      - FEMELLE / MALE / PAIR qualifiers
      - Specimen numbers (#1, #2, etc.)
      - Condition codes (A1, A2)
      - RARE marker
      - Location detail after colon (→ location field, not description)

    Returns (latin_name, description_parts, colon_location_detail).
    """
    if not text or not text.strip():
        return None, [], None

    original = str(text)
    parts = []
    colon_loc = None
    remaining = original

    # ── 1. French location patterns ──────────────────────────────────────
    # These contribute to location, NOT description — skip adding to parts.
    for pattern, _ in FRENCH_LOCATION_PATTERNS:
        remaining = pattern.sub("", remaining)

    # ── 2. Location detail after colon ──────────────────────────────────
    # e.g., "d'Espagne: Albarracin" or "d'Espagne: Granada"
    # This is location detail — capture for the location field, not description
    m = re.search(r':\s*(.+)$', remaining)
    if m:
        after_colon = m.group(1).strip()
        # Only treat as location detail if it contains no gender/condition keywords
        if not any(kw in after_colon.upper() for kw in
                   ["FEMELLE", "MALE", "PAIR", "RARE", "A1", "A2", "ABERRATION"]):
            colon_loc = after_colon
            remaining = remaining[:m.start()].strip()
        else:
            # Contains specimen keywords — keep in remaining to be stripped
            remaining = remaining[:m.start()].strip()

    # ── 3. FEMELLE / MALE / PAIR ────────────────────────────────────────
    for kw in ["FEMELLE", "MALE", "PAIR"]:
        if kw in remaining.upper():
            m = re.search(r'\b' + kw + r'\b\s*\S*', remaining, re.IGNORECASE)
            if m:
                parts.append(m.group(0).strip())
                remaining = remaining[:m.start()] + remaining[m.end():]

    # ── 4. ABERRATION patterns ──────────────────────────────────────────
    m = re.search(r'\bABERRATION\b\s*\S+', remaining, re.IGNORECASE)
    if m:
        parts.append(m.group(0).strip())
        remaining = remaining[:m.start()] + remaining[m.end():]

    # ── 5. Condition codes (A1, A2) ─────────────────────────────────────
    m = re.search(r'\bA[12]\b', remaining)
    if m:
        parts.append(m.group(0))
        remaining = remaining[:m.start()] + remaining[m.end():]

    # ── 6. RARE ──────────────────────────────────────────────────────────
    m = re.search(r'\bRARE\b', remaining)
    if m:
        parts.append("RARE")
        remaining = remaining[:m.start()] + remaining[m.end():]

    # ── 7. Specimen numbers (#1, #2, etc.) ──────────────────────────────
    m = re.search(r'#[0-9]+', remaining)
    if m:
        parts.append(f"specimen {m.group(0)}")
        remaining = remaining[:m.start()] + remaining[m.end():]

    # Clean up multiple spaces
    remaining = re.sub(r'\s+', ' ', remaining).strip()

    # If we stripped everything, use the original
    latin_name = remaining if remaining else original.strip()

    return latin_name, parts, colon_loc


# ══════════════════════════════════════════════════════════════════════════
#  SUPPLIER PARSING
# ══════════════════════════════════════════════════════════════════════════

EMAIL_RE = re.compile(r'[\w.+-]+@[\w.-]+\.[a-zA-Z]{2,}')


def parse_supplier(text):
    """Parse supplier field into username, name, address, and email.

    Format:  \ufeff...'username' Name, Address lines [email@example.com]

    Returns (username, name, address, email).
    """
    if not text or not text.strip():
        return None, None, None, None

    raw = str(text)

    # Strip BOM characters (\ufeff)
    raw = raw.replace("\ufeff", "")
    raw = raw.strip()

    # Extract email before collapsing whitespace
    email_match = EMAIL_RE.search(raw)
    email = email_match.group(0).lower() if email_match else None
    if email_match:
        raw = raw[:email_match.start()] + raw[email_match.end():]

    # Extract username from first pair of single quotes
    m = re.match(r"'([^']+)'\s*", raw)
    username = m.group(1).strip() if m else None
    rest = raw[m.end():] if m else raw

    # Collapse whitespace (including newlines) to single spaces
    rest = re.sub(r'\s+', ' ', rest).strip()

    if not rest:
        return username, None, None, email

    # Split name from address.
    # Strategy: name is the part before the first comma that looks like a person name.
    # If no comma, take the first 1-2 tokens as name.
    comma_idx = rest.find(",")
    if comma_idx > 0 and comma_idx < 50:
        potential_name = rest[:comma_idx].strip()
        address = rest[comma_idx + 1:].strip()

        name_tokens = potential_name.split()
        if len(name_tokens) <= 3:
            name = potential_name
        else:
            name = " ".join(name_tokens[:2])
            address = " ".join(name_tokens[2:]) + ", " + address
    else:
        tokens = rest.split()
        if len(tokens) == 1:
            name = tokens[0]
            address = None
        elif len(tokens) == 2:
            name = " ".join(tokens[:2])
            address = None
        else:
            raw_no_bom = raw[m.end():].strip() if m else raw
            if "\n" in raw_no_bom:
                lines = raw_no_bom.split("\n")
                name = lines[0].strip()
                address = " ".join(l.strip() for l in lines[1:] if l.strip())
            else:
                name = " ".join(tokens[:2])
                address = " ".join(tokens[2:])

    if name:
        name = re.sub(r'\s+', ' ', name).strip().rstrip(",")
    if address:
        address = re.sub(r'\s+', ' ', address).strip().rstrip(",")

    return username, name, address, email


# ══════════════════════════════════════════════════════════════════════════
#  PRICE PARSING
# ══════════════════════════════════════════════════════════════════════════

def parse_price(value):
    """Parse the price column.

    Returns (price_number, is_collected, price_note).
    """
    if value is None:
        return None, False, "missing"

    if isinstance(value, (int, float)):
        return float(value), False, None

    if isinstance(value, str):
        v = value.strip().lower()
        if v == "collected":
            return None, True, None
        elif v in ("free", "fee"):
            return 0.0, False, "freebie"
        elif v == "job lot":
            return None, False, "job lot"
        elif v == "see below":
            return None, False, "see below"
        elif v == "not paid yet":
            return None, False, "not paid yet"
        elif v == "4..49":
            # Likely a typo for 4.49
            return None, False, "unparseable: '4..49'"
        else:
            return None, False, f"unparseable: {value!r}"

    return None, False, f"unexpected type: {type(value).__name__}"


# ══════════════════════════════════════════════════════════════════════════
#  MAIN PARSER
# ══════════════════════════════════════════════════════════════════════════

def parse_specimen(row_idx, row_cells):
    """Parse one spreadsheet row into a structured specimen object.

    Returns (specimen_dict, errors_or_warnings_list).
    """
    issues = []

    # ── Read raw cell values ────────────────────────────────────────────
    notes = str(row_cells[1] or "").strip() if row_cells[1] else ""
    eng_raw = str(row_cells[2] or "").strip() if row_cells[2] else ""
    lat_raw = str(row_cells[3] or "").strip() if row_cells[3] else ""
    sup_raw = str(row_cells[4] or "").strip() if row_cells[4] else ""
    price_raw = row_cells[5]
    date_bought_raw = row_cells[6]
    date_sent_raw = row_cells[7]
    date_received_raw = row_cells[8]
    setting_board = row_cells[9]
    cat_number = row_cells[10]

    if not eng_raw and not lat_raw:
        return None, ["empty row"]

    # Normalize non-breaking space
    eng_raw = eng_raw.replace("\xa0", " ")

    # ── Parse English name column ───────────────────────────────────────
    all_desc_parts = []

    # 1. Sex and quantity
    sex, qty_parts, eng_after_sex, had_pair = extract_sex_and_quantity(eng_raw)
    all_desc_parts.extend(qty_parts)

    # 2. Location ("from X")
    location, eng_after_loc = extract_location(eng_after_sex)

    # 3. Altitude
    altitude_m, eng_after_alt = extract_altitude(eng_after_loc)

    # 4. Description from col C
    desc_c, eng_after_desc = extract_description(eng_after_alt)
    all_desc_parts.extend(desc_c)

    # 5. The remainder is the English name
    english_name = re.sub(r'\s+', ' ', eng_after_desc).strip()
    # Clean trailing punctuation left after stripping sex/location/description
    if english_name:
        # Remove trailing periods, commas, semicolons, colons, question marks
        english_name = english_name.rstrip(".,;:?!")

    # ── If location not found from col C, try extraction from col D ──────
    lat_location, _ = extract_location_from_latin(lat_raw)
    if location is None and lat_location is not None:
        location = lat_location
    elif location is not None and lat_location is not None:
        # Both have location — prefer the more specific one from col C
        pass

    # If english_name still contains the location text as a suffix, strip it.
    # This handles cases like "American Painted Lady Canary Isles" where there
    # was no "from" prefix, but the Latin column revealed the location.
    if location and english_name:
        loc_clean = location.rstrip(".,;")
        name_lower = english_name.lower()
        # Exact suffix match (case-insensitive)
        if name_lower.endswith(loc_clean.lower()):
            english_name = english_name[:-len(loc_clean)].strip()
        else:
            # Try matching last N words
            name_words = english_name.split()
            loc_words = loc_clean.split()
            if len(loc_words) >= 1 and len(name_words) >= len(loc_words):
                suffix = " ".join(name_words[-len(loc_words):])
                if suffix.lower() == loc_clean.lower():
                    english_name = " ".join(name_words[:-len(loc_words)])

    # ── Clean english_name remnants ────────────────────────────────────
    if english_name:
        # Remove trailing "from" with no location
        english_name = re.sub(r'\bfrom\s*$', '', english_name, flags=re.IGNORECASE).strip()

        # Remove trailing " +" and " + u" etc.
        english_name = re.sub(r'\s*\+\s*\S*\s*$', '', english_name).strip()

        # Remove trailing abbreviations: u(unset), m(male), f(female), p(pair), ab(aberration)
        # and with period: u., m., f., ab.
        english_name = re.sub(r'\s+(?:[umfpab]\.?)\s*$', '', english_name, flags=re.IGNORECASE).strip()

        # Remove trailing ? marks
        english_name = english_name.rstrip('?')

        # Remove trailing dots (from "santateresae." etc.)
        english_name = english_name.rstrip('.')

        # Check for leftover form/subspecies names at the end of english_name.
        # Scientific form names are typically single words ending in
        # -is, -ae, -us, -um, -i, -a that appear after the common name.
        # Move them to description.
        words = english_name.split()
        if len(words) >= 2:
            last_word = words[-1]
            # Strip trailing ? and . for matching
            last_clean = last_word.rstrip('?.,;')
            # Check if last word looks like a scientific form name
            # (may be capitalized or lowercase, ends in typical taxonomic suffixes)
            if (len(last_clean) > 3
                    and re.search(r'(?:is|ae|us|um|i|a|ensis|ensis)$', last_clean, re.IGNORECASE)
                    and last_clean.lower() not in {'this', 'from', 'with', 'that', 'they', 'have', 'been', 'their'}):
                # Accept if previous word is capitalized OR is a short token
                # like "x" or a digit (quantity remnants)
                second_last = words[-2] if len(words) >= 2 else ''
                accept = False
                if second_last:
                    if second_last[0].isupper():
                        accept = True
                    elif second_last.lower() in ('x',) or second_last.isdigit():
                        accept = True
                if accept:
                    desc_fragment = last_word
                    all_desc_parts.append(desc_fragment)
                    english_name = " ".join(words[:-1])

    # Remove "N x" quantity remnants that lost their sex indicator
    if english_name:
        english_name = re.sub(r'\b\d+\s*x\s*', '', english_name).strip()

    # Handle "from" in english_name where location was empty
    if english_name:
        english_name = re.sub(r'\s+', ' ', english_name).strip()

    # ── Altitude in Latin column too ────────────────────────────────────
    lat_altitude, _ = extract_altitude(lat_raw)
    if altitude_m is None and lat_altitude is not None:
        altitude_m = lat_altitude

    # ── Clean and parse Latin name ──────────────────────────────────────
    latin_name, desc_lat, colon_loc = clean_latin_name(lat_raw)
    all_desc_parts.extend(desc_lat)

    # Incorporate colon-location detail from Latin name (e.g. "Albarracin")
    if colon_loc:
        if location:
            location = f"{location}, {colon_loc}"
        else:
            location = colon_loc

    # ── Parse supplier ──────────────────────────────────────────────────
    supplier_username, supplier_name, supplier_address, supplier_email = parse_supplier(sup_raw)

    # ── Detect collector ────────────────────────────────────────────────
    collector = None
    if COLLECTOR_PATTERNS.search(sup_raw or ""):
        collector = "E.D. Adamson"

    # ── Parse price ─────────────────────────────────────────────────────
    price_num, is_collected, price_note = parse_price(price_raw)

    # Collector override from price field
    if is_collected:
        collector = "E.D. Adamson"

    # ── Format dates ────────────────────────────────────────────────────
    date_bought = format_date(date_bought_raw)
    date_sent = format_date(date_sent_raw)
    date_received = format_date(date_received_raw)

    # Log unparseable dates
    if date_bought_raw is not None and date_bought is None:
        issues.append(f"unparseable bought date: {date_bought_raw!r}")
    if date_sent_raw is not None and date_sent is None:
        issues.append(f"unparseable sent date: {date_sent_raw!r}")
    if date_received_raw is not None and date_received is None:
        issues.append(f"unparseable received date: {date_received_raw!r}")

    # ── Setting board and cat number ────────────────────────────────────
    if setting_board is not None:
        setting_board = str(setting_board).strip()
        if setting_board == "" or setting_board.lower() == "none":
            setting_board = None
    if cat_number is not None:
        cat_number = str(cat_number).strip()
        if cat_number == "" or cat_number.lower() == "none":
            cat_number = None

    # ── Build description string ────────────────────────────────────────
    # Filter and deduplicate description parts
    seen = set()
    unique_desc = []
    for part in all_desc_parts:
        clean = part.strip()
        if clean and clean.lower() not in seen:
            seen.add(clean.lower())
            unique_desc.append(clean)
    description = ", ".join(unique_desc) if unique_desc else None

    # ── Collector note ──────────────────────────────────────────────────
    # If notes mention collector or price is collected, add to notes
    notes_final = notes if notes else None
    if price_note and price_note not in ("missing", None) and price_note != "freebie":
        notes_final = (notes_final + "; " if notes_final else "") + f"price: {price_note}"

    # ── Build the specimen object ───────────────────────────────────────
    specimen = {
        "id": str(uuid.uuid4()),
        "english_name": english_name or None,
        "sex": sex,
        "latin_name": latin_name,
        "description": description,
        "location": location,
        "altitude_m": altitude_m,
        "supplier_username": supplier_username,
        "supplier_name": supplier_name,
        "supplier_address": supplier_address,
        "supplier_email": supplier_email,
        "price": price_num,
        "price_is_collected": is_collected,
        "currency": "£",
        "date_bought": date_bought,
        "date_sent": date_sent,
        "date_received": date_received,
        "setting_board": setting_board,
        "cat_number": cat_number,
        "collector": collector,
        "photos": [],
        "notes": notes_final,
    }

    return specimen, issues


# ══════════════════════════════════════════════════════════════════════════
#  REPORT COLLECTION
# ══════════════════════════════════════════════════════════════════════════

def write_report(report_rows, filepath):
    """Write the import report to a text file."""
    lines = []
    lines.append("=" * 72)
    lines.append("BUTTERFLY COLLECTION IMPORT REPORT")
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 72)
    lines.append("")

    if not report_rows:
        lines.append("No parsing issues found — all rows imported cleanly.")
    else:
        lines.append(f"Total rows with issues: {len(report_rows)}")
        lines.append("")
        for row_idx, issues in report_rows:
            lines.append(f"── Row {row_idx} ──────────────────────────────────")
            for issue in issues:
                lines.append(f"  • {issue}")
            lines.append("")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    return len(report_rows)


# ══════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("BUTTERFLY COLLECTION IMPORTER")
    print(f"Input: {INPUT_FILE}")
    print("=" * 60)

    # ── Ensure output directory exists ──────────────────────────────────
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Load workbook ───────────────────────────────────────────────────
    print(f"\n📂 Loading workbook...")
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    print(f"   Sheet: {SHEET_NAME}, Rows: {ws.max_row}, Cols: {ws.max_column}")

    # ── Iterate rows ────────────────────────────────────────────────────
    specimens = []
    report_rows = []
    empty_count = 0
    error_count = 0

    print(f"\n🔍 Parsing {ws.max_row - 1} data rows...")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        has_data = any(cell is not None for cell in row)
        if not has_data:
            empty_count += 1
            continue

        # Skip rows that have absolutely nothing in columns C and D
        c_val = row[2] if len(row) > 2 else None
        d_val = row[3] if len(row) > 3 else None
        if not c_val and not d_val:
            empty_count += 1
            continue

        specimen, issues = parse_specimen(row_idx, row)

        if specimen is None:
            empty_count += 1
            continue

        specimens.append(specimen)

        if issues:
            error_count += 1
            report_rows.append((row_idx, issues))

        if row_idx % 200 == 0:
            print(f"   Processed row {row_idx}... ({len(specimens)} specimens so far)")

    wb.close()

    # ── Write output JSON ───────────────────────────────────────────────
    print(f"\n💾 Writing {len(specimens)} specimens to {OUTPUT_JSON}...")
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(specimens, f, indent=2, ensure_ascii=False)
    json_size = os.path.getsize(OUTPUT_JSON)
    print(f"   ✅ JSON file size: {json_size:,} bytes")

    # ── Write report ────────────────────────────────────────────────────
    report_count = write_report(report_rows, OUTPUT_REPORT)
    print(f"📄 Import report written to {OUTPUT_REPORT}")
    print(f"   Rows with issues: {report_count}")

    # ── Summary ─────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print("IMPORT SUMMARY")
    print("=" * 60)
    print(f"   Total rows in sheet:     {ws.max_row - 1}")
    print(f"   Empty/skipped rows:      {empty_count}")
    print(f"   Specimens parsed:        {len(specimens)}")
    print(f"   Rows in import report:   {report_count}")
    print(f"   Clean imports:           {len(specimens) - error_count}")

    # Compute some stats
    with_sex = sum(1 for s in specimens if s["sex"])
    with_location = sum(1 for s in specimens if s["location"])
    with_altitude = sum(1 for s in specimens if s["altitude_m"] is not None)
    with_collector = sum(1 for s in specimens if s["collector"])
    with_price = sum(1 for s in specimens if s["price"] is not None)
    collected = sum(1 for s in specimens if s["price_is_collected"])

    print(f"   With sex:                {with_sex}")
    print(f"   With location:           {with_location}")
    print(f"   With altitude:           {with_altitude}")
    print(f"   With collector:          {with_collector}")
    print(f"   With numeric price:      {with_price}")
    print(f"   Collected (self):        {collected}")

    # Sample first 3
    print("\n" + "=" * 60)
    print("SAMPLE — First 3 specimens:")
    print("=" * 60)
    for i, s in enumerate(specimens[:3]):
        print(f"\n--- Specimen {i + 1} ---")
        print(json.dumps(s, indent=2, ensure_ascii=False))

    print("\n✅ Import complete!")


if __name__ == "__main__":
    main()
